#!/usr/bin/env python3
"""
Extract data from the metacarpal repair outcomes Excel file and generate
a ready-to-paste src/data.js export block.

Usage:
    pip3 install -r requirements.txt   # one-time setup
    python3 scripts/extract_data.py

Or via npm:
    npm run extract
"""

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_GLOB = list(REPO_ROOT.glob("*.xlsx"))

if not XLSX_GLOB:
    sys.exit("ERROR: No .xlsx file found in repo root.")

XLSX_PATH = sorted(XLSX_GLOB)[0]
print(f"Reading: {XLSX_PATH.name}\n{'='*60}")

try:
    import openpyxl
except ImportError:
    sys.exit(
        "ERROR: openpyxl not installed.\n"
        "Run:  pip3 install -r requirements.txt\n"
        "Then re-run this script."
    )

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ── dump raw sheet contents ──────────────────────────────────────────────────
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)
            if any(c is not None for c in row)]

    print(f"\n── Sheet: {sheet_name!r} ({len(rows)} non-empty rows) ──")
    for i, row in enumerate(rows):
        print(f"  row {i:>3}: {row}")

    # ── parse: row 0 = title, row 1 = column headers, rows 2+ = data ────────
    if len(rows) < 3:
        print("  (too few rows to parse)")
        continue

    title_row = rows[0]
    header_row = rows[1]
    data_rows = rows[2:]

    title = next((str(c).strip() for c in title_row if c is not None), sheet_name)
    headers = [str(h).strip() if h is not None else f"col{i}"
               for i, h in enumerate(header_row)]

    print(f"\n  Title : {title}")
    print(f"  Headers: {headers}")

    # ── emit computed rates ──────────────────────────────────────────────────
    GROUP_COL = headers[0]   # first column = comorbidity/group label
    N_COL = headers[1]       # second column = total patients
    OUTCOME_COLS = headers[2:]

    print(f"\n  Computed rates (%) per outcome:\n")
    rate_header = f"{'Group':<32} {'N':>7}  " + "  ".join(f"{h[:12]:>12}" for h in OUTCOME_COLS)
    print(f"  {rate_header}")
    print(f"  {'-'*len(rate_header)}")

    for row in data_rows:
        d = dict(zip(headers, row))
        group = str(d.get(GROUP_COL, "")).strip()
        n = d.get(N_COL)
        if n is None or n == 0:
            continue
        rates = []
        for col in OUTCOME_COLS:
            v = d.get(col)
            rate = round(float(v) / float(n) * 100, 2) if v is not None else None
            rates.append(rate)
        rate_str = "  ".join(f"{(r if r is not None else 'N/A'):>12}" for r in rates)
        print(f"  {group:<32} {int(n):>7}  {rate_str}")

    # ── generate JS export ───────────────────────────────────────────────────
    print(f"\n\n{'='*60}")
    print(f"// GENERATED: export for sheet '{sheet_name}'")
    print(f"// Paste into src/data.js")
    print(f"{'='*60}\n")

    js_name = "metacarpalComorbidities"
    print(f"export const {js_name} = [")
    for row in data_rows:
        d = dict(zip(headers, row))
        group = str(d.get(GROUP_COL, "")).strip()
        n = d.get(N_COL)
        if n is None:
            continue
        n_val = int(float(n))

        outcomes = {}
        for col in OUTCOME_COLS:
            v = d.get(col)
            count = int(float(v)) if v is not None else 0
            safe_key = (col.lower()
                        .replace(" ", "_")
                        .replace(",", "")
                        .replace("+", "plus")
                        .replace("-", "_")
                        .replace(".", "")
                        .replace("/", "_"))
            outcomes[safe_key] = count

        # build JS object
        outcome_parts = ", ".join(f"{k}: {v}" for k, v in outcomes.items())
        print(f'  {{ group: "{group}", n: {n_val}, {outcome_parts} }},')

    print("];")
